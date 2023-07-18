from openpyxl import Workbook, load_workbook
import pandas as pd


def wellness_question_data():
  root_path = "/content/drive/MyDrive/Project/goorm_project/goorm_project_3/data/"
  wellness_file = root_path + "웰니스_대화_스크립트_데이터셋.xlsx"
  wellness_q_output = root_path + "wellness_dialog_question.txt"

  f = open(wellness_q_output, 'w')

  wb = load_workbook(filename=wellness_file)

  ws = wb[wb.sheetnames[0]]
  # print(sheet)
  for row in ws.iter_rows():
    f.write(row[0].value + "    " + row[1].value + "\n")

  f.close()

def wellness_answer_data():
  root_path = "/content/drive/MyDrive/Project/goorm_project/goorm_project_3/data/"
  wellness_file = root_path + "웰니스_대화_스크립트_데이터셋.xlsx"
  wellness_a_output = root_path + "wellness_dialog_answer.txt"

  f = open(wellness_a_output, 'w')
  wb = load_workbook(filename=wellness_file)
  ws = wb[wb.sheetnames[0]]

  for row in ws.iter_rows():
    if row[2].value == None:
      continue
    else:
      f.write(row[0].value + "    " + row[2].value + "\n")
  f.close()

def wellness_dialog_for_autoregressive():
  root_path = "/content/drive/MyDrive/Project/goorm_project/goorm_project_3/data/"
  # wellness_file = root_path + "/wellness_dialog_dataset.xlsx"
  wellness_answer_file = root_path + "wellness_dialog_answer.txt"
  wellness_question_file = root_path + "wellness_dialog_question.txt"
  wellness_autoregressive_file = root_path + "wellness_dialog_for_autoregressive.txt"


  answ_file = open(wellness_answer_file, 'r')
  ques_file = open(wellness_question_file, 'r')
  autoregressive_file = open(wellness_autoregressive_file, 'w')

  answ_lines = answ_file.readlines()
  ques_lines = ques_file.readlines()
  ques_dict = {}
  for line_num, line_data in enumerate(ques_lines):
    ques_data = line_data.split('    ')
    for ans_line_num, ans_line_data in enumerate(answ_lines):
      ans_data = ans_line_data.split('    ')
      if ques_data[0] == ans_data[0]:
        autoregressive_file.write(ques_data[1][:-1] + "    " + ans_data[1])
      else:
        continue

  answ_file.close()
  ques_file.close()
  autoregressive_file.close()

def living_dialog_data():
  root_path = "/content/drive/MyDrive/Project/goorm_project/goorm_project_3/data/"
  wellness_autoregressive_file = root_path + "wellness_dialog_for_autoregressive.txt"
  living_file = root_path + "ChatbotData.txt"
  wellness_output = root_path + "dialog.txt"

  f = open(wellness_output, 'w')
  wellness_data = open(wellness_autoregressive_file, 'r')
  living_data = open(living_file, 'r')

  wellness_lines = wellness_data.readlines()
  living_lines = living_data.readlines()

  for idx, row in enumerate(wellness_lines):
    if idx == 0:
      continue
    f.write(row)
  for idx, row in enumerate(living_lines):
    if idx == 0:
      continue
    f.write(row[:-3] + "\n")
    

  f.close()

wellness_question_data()
wellness_answer_data()
wellness_dialog_for_autoregressive()
living_dialog_data()